# Standard Library Imports
import tarfile
import time
import json
import argparse
import re
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

INPUT_FILE_FORMAT = "json"
LOGFILE = "config_analysis_log.log"
DEBUG = True
VERBOSE_OUTPUT = True


class OutputLogging():
    """
    Output Logging
    """

    def __init__(self, file_name):
        now = time.time()
        local_time = time.localtime(now)
        timestamp = time.strftime("%Y-%m-%d-%H%M", local_time)
        self.output_file = timestamp + "-" + file_name
        self.output = open(self.output_file, "a")

    def print_string_to_file(self, message):
        message = message + "\n"
        self.output.write(message)

    def print_to_stdout(self, message):
        print message

    def print_message(self, message):
        """ Print message to both STDOUT and output Logfile"""
        # self.print_to_stdout(message)
        self.print_string_to_file(message)


LOG = OutputLogging(LOGFILE)


class AnalysisOutput():
    """
    Used to save the output of the Analysis to an Excel sheet
    """

    def __init__(self, file_name):
        self.file_name = file_name
        self.wb = Workbook()

        '''
        Define and create Worksheets
        Each dict entry has the following format ['sheet_object', 'row_count', 'last column'
        '''
        self.wb_sheets = {
            'domain': [self.wb.create_sheet("phys_domain"), 1, 'A'],
            'tenant': [self.wb.create_sheet("tenant"), 1, 'A']
        }

        ''' Phys domain Sheet Header '''
        self.wb_sheets['domain'][0].append(['name', 'type', 'vlan_pool'])
        self.wb_sheets['domain'][1] = 1
        self.wb_sheets['domain'][2] = 'C'

        ''' Tenant Sheet Header '''
        self.wb_sheets['tenant'][0].append(
            ['name', 'description', 'ownerKey', 'ownerTag'])
        self.wb_sheets['tenant'][1] = 1
        self.wb_sheets['tenant'][2] = 'D'

    def add_row(self, sheet, data):
        self.wb_sheets[sheet][0].append(data)
        self.wb_sheets[sheet][1] = self.wb_sheets[sheet][1] + 1

    def save_to_disk(self):
        style = TableStyleInfo(name="TableStyleLight9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)

        ''' Format Sheet Tables '''
        for sheet in self.wb_sheets.keys():
            tab = Table(displayName=sheet, ref="A1:%s%s" %
                        (self.wb_sheets[sheet][2], self.wb_sheets[sheet][1]))
            tab.tableStyleInfo = style
            self.wb_sheets[sheet][0].add_table(tab)

        ''' Save Wordbook to disk '''
        self.wb.save(self.file_name)


ANALYSIS_OUTPUT = AnalysisOutput("analysis.xlsx")


class ConfigAnalyse():
    """ 
    Takes the name/path to the tgz backup file : input_file
    The format of the config backup file : input_format
    """

    def __init__(self, input_file, input_format):
        self.input_file = input_file

        if input_format == "json":
            LOG.print_message("+- Starting Configuration Analysis....")
        else:
            LOG.print_message("+-- Only JSON input are supported")

        self.config_data = self.read_config_backup(input_file)
        self.analyse_config(self.config_data)

    def read_config_backup(self, input_file):
        config_data = dict(fvTenant={},
                           physDomP={}
                           )

        ''' Open TAR file, and raise an exception if it fails '''
        try:
            LOG.print_message("+- Reading Config Archive %s" % input_file)
            tar = tarfile.open(input_file, "r:gz")
        except:
            raise

        for tarinfo in tar:
            ''' Read the TAR file, and skip entries that are not regular files or is of 0 bytes in size '''
            if tarinfo.isreg() and tarinfo.size > 0:
                ''' Read file and add json information to json_content dict'''
                if DEBUG:
                    LOG.print_message("+-- Reading file %s" % tarinfo.name)
                file = tar.extractfile(tarinfo.name)
                content = file.read()

                ''' Attempt to decode file as json '''
                try:
                    json_content = json.loads(content)

                    ''' Analyse only json_content belonging to polUni '''
                    if 'polUni' in json_content.keys():
                        for policy_root in json_content['polUni']['children']:
                            for policy_class in policy_root:
                                if policy_class == 'fvTenant':
                                    config_data['fvTenant'][policy_root[policy_class]
                                                            ['attributes']['name']] = policy_root[policy_class]
                                elif policy_class == 'physDomP':
                                    config_data['physDomP'][policy_root[policy_class]
                                                            ['attributes']['name']] = policy_root[policy_class]
                                elif policy_class == 'quotaCont' or policy_class == 'plannerCont' or policy_class == 'aaaRbacEp' or policy_class == 'dbgDebugP' or policy_class == 'pkiFabricCommunicationEp':
                                    ''' Unsupported Policies '''
                                    if DEBUG:
                                        LOG.print_message(
                                            "+--- Skipping Policy: polUni/%s - Not supported" % policy_class)
                                else:
                                    ''' Skipping unsupported polUni objects '''
                                    if DEBUG:
                                        LOG.print_message(
                                            "+--- Skipping Config: polUni/%s - Unknown Policy" % policy_class)
                                        #print json.dumps(policy_root[policy_class], indent=4, sort_keys=True)
                    elif 'topRoot' in json_content.keys():
                        ''' Skipping topRoot as this one does not contain configuration items'''
                        if DEBUG:
                            LOG.print_message(
                                "+--- Skipping policy root: polUni/%s - Not supported" % policy_class)

                        skip = True
                    else:
                        ''' Skipping unsupported root objects '''
                        LOG.print_message(
                            "+--- Skipping Config: %s - Unknown Policy Root" % str(json_content.keys()[0]))
                except ValueError:
                    if DEBUG:
                        LOG.print_message(
                            "+--- Skipping, as file does not contain JSON data")
        tar.close()
        return(config_data)

    def analyse_config(self, config_data):
        LOG.print_message("+- Starting Configuration Analysis")

        for key in config_data.keys():
            if key == 'fvTenant':
                LOG.print_message("+-- Analysing Tenant Configuration")

                for tenant in config_data[key]:
                    self.analyse_tenant(config_data[key][tenant])
            elif key == 'physDomP':
                LOG.print_message(
                    "+-- Analysing Physical Domain Configuration")

                for domain in config_data[key]:
                    self.analyse_physdom(config_data[key][domain])
            else:
                LOG.print_message(
                    "+-- WARNING, Unsupported class found: %s" % key)
                raise Exception
            #print config_data[key]

        LOG.print_message(("+- Saving Analysis Output to disk"))
        ANALYSIS_OUTPUT.save_to_disk()

    def analyse_tenant(self, data):
        LOG.print_message("+--- Tenant Found: %s" % data['attributes']['name'])
        tenant_name = data['attributes']['name']

        '''' Tenant Attributes '''
        if VERBOSE_OUTPUT:
            LOG.print_message("+---- descr: %s" % data['attributes']['descr'])
            LOG.print_message("+---- dn: %s" % data['attributes']['dn'])
            LOG.print_message("+---- name: %s" % data['attributes']['name'])
            LOG.print_message("+---- nameAlias: %s" %
                              data['attributes']['nameAlias'])
            LOG.print_message("+---- ownerKey: %s" %
                              data['attributes']['ownerKey'])
            LOG.print_message("+---- ownerTag: %s" %
                              data['attributes']['ownerTag'])

        ''' Add Found Tenant to Analysis Output '''
        ANALYSIS_OUTPUT.add_row('tenant', [data['attributes']['name'], data['attributes']
                                           ['descr'], data['attributes']['ownerKey'], data['attributes']['ownerTag']])

    def analyse_physdom(self, data):
        domain_name = data['attributes']['name']
        LOG.print_message("+--- Physical Domain Found: %s" %
                          data['attributes']['name'])

        ''' VRF Object Attributes '''
        if VERBOSE_OUTPUT:
            LOG.print_message("+------ dn: %s" % data['attributes']['dn'])
            LOG.print_message("+------ name: %s" % data['attributes']['name'])
            LOG.print_message("+------ nameAlias: %s" %
                              data['attributes']['nameAlias'])
            LOG.print_message("+------ ownerKey: %s" %
                              data['attributes']['ownerKey'])
            LOG.print_message("+------ ownerTag: %s" %
                              data['attributes']['ownerTag'])

        ''' Object Children'''
        try:
            for child in data['children']:
                for child_class in sorted(child.keys()):
                    if child_class == 'infraRsVlanNs':
                        ''' infraRsVlanNs Child Object Attributes '''
                        LOG.print_message(
                            "+---- Associated VLAN Pool Found: %s" % child[child_class]['attributes']['tDn'])

                        reg_result = re.search(
                            'uni\/infra\/vlanns-\[(.+)\]\-(static|dynamic)', child[child_class]['attributes']['tDn'])
                        phys_domain = reg_result.group(1)

                        ANALYSIS_OUTPUT.add_row(
                            'domain', [domain_name, 'physical', phys_domain])

                        if VERBOSE_OUTPUT:
                            LOG.print_message(
                                "+------ Associated VLAN Pool Configuration")
                            LOG.print_message("+------- dn: %s" %
                                              child[child_class]['attributes']['dn'])
                            LOG.print_message(
                                "+------- tDn: %s" % child[child_class]['attributes']['tDn'])
        except KeyError:
            ''' No children objects '''
            if DEBUG:
                LOG.print_message("+------ WARNING, No physDomP child objects")


''' Grab INPUT_FILE from command arguments and perform basic validation '''
parser = argparse.ArgumentParser(
    description='''This scrips analysis configuaration backup file in JSON format,
    and output is saved in an Excel wordbook''')
parser.add_argument('-i', '--input', help='Input file name',
                    dest="input", required=True)
args = parser.parse_args()
cmd_parameters = vars(args)

if "tar.gz" in cmd_parameters['input']:
    ConfigAnalyse(cmd_parameters['input'], INPUT_FILE_FORMAT)
else:
    print "ERROR: Input file name does not seem to be a tar.gz file"
